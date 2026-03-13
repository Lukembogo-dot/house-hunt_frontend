"""
Manual Test Documentation Generator
Generates an Excel file with manual test case documentation for House Hunt QA
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


def create_manual_test_excel(filename="Manual_Test_Documentation.xlsx"):
    """
    Creates an Excel file with manual test documentation columns.
    
    Args:
        filename: Name of the Excel file to create
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Manual Test Cases"
    
    # Define the column headers
    headers = [
        "Test Scenario ID",
        "Test Scenario Description",
        "Test Case ID",
        "Test Case Description",
        "Test Steps",
        "Preconditions",
        "Test Data",
        "Post Conditions",
        "Expected Result",
        "Actual Result",
        "Status(Pass, Fail, Blocked, Not Executed)",
        "Executed By",
        "Executed Date",
        "Comments (if any)"
    ]
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths for better readability
    column_widths = {
        'A': 18,  # Test Scenario ID
        'B': 35,  # Test Scenario Description
        'C': 15,  # Test Case ID
        'D': 40,  # Test Case Description
        'E': 50,  # Test Steps
        'F': 35,  # Preconditions
        'G': 30,  # Test Data
        'H': 35,  # Post Conditions
        'I': 40,  # Expected Result
        'J': 25,  # Actual Result
        'K': 20,  # Status
        'L': 18,  # Executed By
        'M': 15,  # Executed Date
        'N': 35   # Comments
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Save the workbook
    wb.save(filename)
    print(f"Excel file '{filename}' created successfully!")
    return filename


def add_test_case(filename, test_data):
    """
    Adds a test case to the existing Excel file.
    
    Args:
        filename: Name of the Excel file
        test_data: Dictionary containing test case data
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Find the next empty row
    next_row = ws.max_row + 1
    
    # Map test data to columns
    data_mapping = {
        'Test Scenario ID': test_data.get('test_scenario_id', ''),
        'Test Scenario Description': test_data.get('test_scenario_description', ''),
        'Test Case ID': test_data.get('test_case_id', ''),
        'Test Case Description': test_data.get('test_case_description', ''),
        'Test Steps': test_data.get('test_steps', ''),
        'Preconditions': test_data.get('preconditions', ''),
        'Test Data': test_data.get('test_data', ''),
        'Post Conditions': test_data.get('post_conditions', ''),
        'Expected Result': test_data.get('expected_result', ''),
        'Actual Result': test_data.get('actual_result', ''),
        'Status(Pass, Fail, Blocked, Not Executed)': test_data.get('status', 'Not Executed'),
        'Executed By': test_data.get('executed_by', ''),
        'Executed Date': test_data.get('executed_date', ''),
        'Comments (if any)': test_data.get('comments', '')
    }
    
    # Get headers to map columns
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    
    # Write test data
    for col, header in enumerate(headers, 1):
        if header in data_mapping:
            ws.cell(row=next_row, column=col, value=data_mapping[header])
    
    wb.save(filename)
    print(f"Test case '{test_data.get('test_case_id', 'Unknown')}' added to '{filename}'")


def create_house_hunt_test_cases(filename):
    """
    Creates comprehensive test cases for House Hunt application.
    
    Args:
        filename: Name of the Excel file
    """
    
    # ==============================================================================
    # 1. USER AUTHENTICATION - LOGIN TESTS
    # ==============================================================================
    
    login_test_cases = [
        {
            'test_scenario_id': 'TS_AUTH_001',
            'test_scenario_description': 'User Login Functionality',
            'test_case_id': 'TC_AUTH_001',
            'test_case_description': 'Verify login page loads with correct heading',
            'test_steps': '1. Navigate to login page (/login)\n2. Check for "Welcome Back" heading',
            'preconditions': 'User is on the login page URL',
            'test_data': 'URL: http://localhost:5173/login',
            'post_conditions': 'User remains on login page',
            'expected_result': 'Page displays "Welcome Back" heading',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_001',
            'test_scenario_description': 'User Login Functionality',
            'test_case_id': 'TC_AUTH_002',
            'test_case_description': 'Verify email/phone input field is visible on login page',
            'test_steps': '1. Navigate to login page\n2. Locate the email/phone input field',
            'preconditions': 'User is on login page',
            'test_data': 'None',
            'post_conditions': 'User remains on login page',
            'expected_result': 'Email or WhatsApp Number input field is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_001',
            'test_scenario_description': 'User Login Functionality',
            'test_case_id': 'TC_AUTH_003',
            'test_case_description': 'Verify Continue button is present on login page',
            'test_steps': '1. Navigate to login page\n2. Locate Continue button',
            'preconditions': 'User is on login page',
            'test_data': 'None',
            'post_conditions': 'User remains on login page',
            'expected_result': 'Continue button is visible and clickable',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_001',
            'test_scenario_description': 'User Login Functionality',
            'test_case_id': 'TC_AUTH_004',
            'test_case_description': 'Verify Google Sign-In option is available',
            'test_steps': '1. Navigate to login page\n2. Check for Google Sign-In option or "Or continue with" text',
            'preconditions': 'User is on login page',
            'test_data': 'None',
            'post_conditions': 'User remains on login page',
            'expected_result': 'Google Sign-In option is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_001',
            'test_scenario_description': 'User Login Functionality',
            'test_case_id': 'TC_AUTH_005',
            'test_case_description': 'Verify link to registration page exists',
            'test_steps': '1. Navigate to login page\n2. Look for "Create an account" link',
            'preconditions': 'User is on login page',
            'test_data': 'None',
            'post_conditions': 'User remains on login page',
            'expected_result': '"Create an account" link is visible with href="/register"',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_001',
            'test_scenario_description': 'User Login Functionality',
            'test_case_id': 'TC_AUTH_006',
            'test_case_description': 'Verify clicking Create an account navigates to registration page',
            'test_steps': '1. Navigate to login page\n2. Click on "Create an account" link\n3. Verify URL changes to /register',
            'preconditions': 'User is on login page',
            'test_data': 'None',
            'post_conditions': 'User is navigated to registration page',
            'expected_result': 'URL contains "/register" and heading "Create your account" is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 2. USER REGISTRATION TESTS
    # ==============================================================================
    
    registration_test_cases = [
        {
            'test_scenario_id': 'TS_AUTH_002',
            'test_scenario_description': 'User Registration Functionality',
            'test_case_id': 'TC_AUTH_007',
            'test_case_description': 'Verify registration page loads with correct heading',
            'test_steps': '1. Navigate to registration page (/register)\n2. Check for "Create your account" heading',
            'preconditions': 'User is on the registration page URL',
            'test_data': 'URL: http://localhost:5173/register',
            'post_conditions': 'User remains on registration page',
            'expected_result': 'Page displays "Create your account" heading',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_002',
            'test_scenario_description': 'User Registration Functionality',
            'test_case_id': 'TC_AUTH_008',
            'test_case_description': 'Verify all required form fields are present',
            'test_steps': '1. Navigate to registration page\n2. Check for: Full Name, Email address, Password, Confirm Password fields',
            'preconditions': 'User is on registration page',
            'test_data': 'None',
            'post_conditions': 'User remains on registration page',
            'expected_result': 'All form fields (Full Name, Email, Password, Confirm Password) are visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_002',
            'test_scenario_description': 'User Registration Functionality',
            'test_case_id': 'TC_AUTH_009',
            'test_case_description': 'Verify terms and conditions checkbox is present',
            'test_steps': '1. Navigate to registration page\n2. Look for terms checkbox with text "I accept the"',
            'preconditions': 'User is on registration page',
            'test_data': 'None',
            'post_conditions': 'User remains on registration page',
            'expected_result': 'Terms checkbox is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_002',
            'test_scenario_description': 'User Registration Functionality',
            'test_case_id': 'TC_AUTH_010',
            'test_case_description': 'Verify Sign up button is present on registration page',
            'test_steps': '1. Navigate to registration page\n2. Look for "Sign up" button',
            'preconditions': 'User is on registration page',
            'test_data': 'None',
            'post_conditions': 'User remains on registration page',
            'expected_result': 'Sign up button (not Google Sign up) is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_002',
            'test_scenario_description': 'User Registration Functionality',
            'test_case_id': 'TC_AUTH_011',
            'test_case_description': 'Verify Google Sign-Up option is available on registration',
            'test_steps': '1. Navigate to registration page\n2. Check for "Or sign up with" text or Google option',
            'preconditions': 'User is on registration page',
            'test_data': 'None',
            'post_conditions': 'User remains on registration page',
            'expected_result': 'Google Sign-Up option is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_002',
            'test_scenario_description': 'User Registration Functionality',
            'test_case_id': 'TC_AUTH_012',
            'test_case_description': 'Verify link to login page exists on registration',
            'test_steps': '1. Navigate to registration page\n2. Look for "Already have an account? Sign in" link',
            'preconditions': 'User is on registration page',
            'test_data': 'None',
            'post_conditions': 'User remains on registration page',
            'expected_result': '"Already have an account? Sign in" link is visible with href="/login"',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_002',
            'test_scenario_description': 'User Registration Functionality',
            'test_case_id': 'TC_AUTH_013',
            'test_case_description': 'Verify error message when passwords do not match',
            'test_steps': '1. Navigate to registration page\n2. Fill Full Name with "Test User"\n3. Fill Email with "test@example.com"\n4. Fill Password with "password123"\n5. Fill Confirm Password with "differentpassword"\n6. Check terms checkbox\n7. Click Sign up button',
            'preconditions': 'User is on registration page',
            'test_data': 'Password: password123, Confirm Password: differentpassword',
            'post_conditions': 'User remains on registration page',
            'expected_result': 'Error message "Passwords do not match." is displayed',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AUTH_002',
            'test_scenario_description': 'User Registration Functionality',
            'test_case_id': 'TC_AUTH_014',
            'test_case_description': 'Verify navigation to login page from registration',
            'test_steps': '1. Navigate to registration page\n2. Click on "Already have an account? Sign in" link',
            'preconditions': 'User is on registration page',
            'test_data': 'None',
            'post_conditions': 'User is navigated to login page',
            'expected_result': 'URL contains "/login" and heading "Welcome Back" is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 3. HOME PAGE - HERO SECTION TESTS
    # ==============================================================================
    
    home_page_test_cases = [
        {
            'test_scenario_id': 'TS_HOME_001',
            'test_scenario_description': 'Home Page - Hero Section',
            'test_case_id': 'TC_HOME_001',
            'test_case_description': 'Verify home page loads with correct title',
            'test_steps': '1. Navigate to home page (http://localhost:5173/)\n2. Check for "Prime Land Opportunities" heading',
            'preconditions': 'User is on the home page',
            'test_data': 'URL: http://localhost:5173/',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Page displays "Prime Land Opportunities" heading',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_HOME_001',
            'test_scenario_description': 'Home Page - Hero Section',
            'test_case_id': 'TC_HOME_002',
            'test_case_description': 'Verify location search textbox is present',
            'test_steps': '1. Navigate to home page\n2. Locate the search textbox with placeholder "Enter Location..."',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Location search textbox is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_HOME_001',
            'test_scenario_description': 'Home Page - Hero Section',
            'test_case_id': 'TC_HOME_003',
            'test_case_description': 'Verify Rent Property button is present',
            'test_steps': '1. Navigate to home page\n2. Look for "Rent Property" button',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': '"Rent Property" button is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_HOME_001',
            'test_scenario_description': 'Home Page - Hero Section',
            'test_case_id': 'TC_HOME_004',
            'test_case_description': 'Verify Filters button is present',
            'test_steps': '1. Navigate to home page\n2. Look for "Filters" button',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': '"Filters" button is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_HOME_001',
            'test_scenario_description': 'Home Page - Hero Section',
            'test_case_id': 'TC_HOME_005',
            'test_case_description': 'Verify search results page loads with heading',
            'test_steps': '1. Navigate to home page\n2. Click on "Rent Property" button\n3. Check for results heading',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User is navigated to search results',
            'expected_result': 'Heading "Land, Apartments & Rentals in" is displayed',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 4. PROPERTY SEARCH & FILTERS TESTS
    # ==============================================================================
    
    property_search_test_cases = [
        {
            'test_scenario_id': 'TS_SEARCH_001',
            'test_scenario_description': 'Property Search Functionality',
            'test_case_id': 'TC_SEARCH_001',
            'test_case_description': 'Verify location search returns relevant results',
            'test_steps': '1. Navigate to home page\n2. Click on location search textbox\n3. Enter a location (e.g., "Kilimani")\n4. Click search button',
            'preconditions': 'User is on the home page with search functionality',
            'test_data': 'Location: Kilimani',
            'post_conditions': 'User is shown search results',
            'expected_result': 'Search results relevant to the entered location are displayed',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_SEARCH_001',
            'test_scenario_description': 'Property Search Functionality',
            'test_case_id': 'TC_SEARCH_002',
            'test_case_description': 'Verify property type filters work correctly',
            'test_steps': '1. Navigate to home page\n2. Click on "Rent Property" button\n3. Select different property types\n4. Verify results update',
            'preconditions': 'User is on property search results',
            'test_data': 'Property types: Rent, Buy, Land',
            'post_conditions': 'Results reflect selected filter',
            'expected_result': 'Properties matching the selected type are displayed',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_SEARCH_001',
            'test_scenario_description': 'Property Search Functionality',
            'test_case_id': 'TC_SEARCH_003',
            'test_case_description': 'Verify Filters button opens filter options',
            'test_steps': '1. Navigate to home page\n2. Click on "Rent Property" button\n3. Click on "Filters" button',
            'preconditions': 'User is on property search results',
            'test_data': 'None',
            'post_conditions': 'Filter panel opens',
            'expected_result': 'Filter options panel is displayed',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_SEARCH_001',
            'test_scenario_description': 'Property Search Functionality',
            'test_case_id': 'TC_SEARCH_004',
            'test_case_description': 'Verify unit type filter in filters panel',
            'test_steps': '1. Navigate to home page\n2. Click on "Rent Property" button\n3. Click on "Filters" button\n4. Locate Unit Type dropdown\n5. Select "1 Bedroom"',
            'preconditions': 'User is on property search results with filters open',
            'test_data': 'Unit Type: 1 Bedroom',
            'post_conditions': 'Results update with selected unit type',
            'expected_result': 'Only 1 Bedroom properties are displayed',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_SEARCH_001',
            'test_scenario_description': 'Property Search Functionality',
            'test_case_id': 'TC_SEARCH_005',
            'test_case_description': 'Verify budget filter in filters panel',
            'test_steps': '1. Navigate to home page\n2. Click on "Rent Property" button\n3. Click on "Filters" button\n4. Locate Budget dropdown\n5. Select "Below 10k"',
            'preconditions': 'User is on property search results with filters open',
            'test_data': 'Budget: Below 10k',
            'post_conditions': 'Results update with selected budget',
            'expected_result': 'Only properties within the selected budget are displayed',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 5. PROPERTY LISTING & DETAILS TESTS
    # ==============================================================================
    
    property_listing_test_cases = [
        {
            'test_scenario_id': 'TS_PROP_001',
            'test_scenario_description': 'Property Listing Display',
            'test_case_id': 'TC_PROP_001',
            'test_case_description': 'Verify featured properties section is displayed',
            'test_steps': '1. Navigate to home page\n2. Scroll to "Hand-Picked Homes We Love" section',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': '"Hand-Picked Homes We Love" section is visible with property listings',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_PROP_001',
            'test_scenario_description': 'Property Listing Display',
            'test_case_id': 'TC_PROP_002',
            'test_case_description': 'Verify property card displays key information',
            'test_steps': '1. Navigate to home page\n2. Click on "Rent Property" button\n3. View property cards',
            'preconditions': 'User is on property search results',
            'test_data': 'None',
            'post_conditions': 'User remains on search results',
            'expected_result': 'Property cards show property name, location, price, and listing type',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_PROP_001',
            'test_scenario_description': 'Property Listing Display',
            'test_case_id': 'TC_PROP_003',
            'test_case_description': 'Verify clicking on property navigates to details',
            'test_steps': '1. Navigate to property search results\n2. Click on a property card',
            'preconditions': 'User is on property search results with properties displayed',
            'test_data': 'None',
            'post_conditions': 'User is navigated to property details page',
            'expected_result': 'Property details page is displayed with full information',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 6. HOUSE HUNT REQUEST FORM TESTS
    # ==============================================================================
    
    house_hunt_request_test_cases = [
        {
            'test_scenario_id': 'TS_HHREQ_001',
            'test_scenario_description': 'House Hunt Request Form',
            'test_case_id': 'TC_HHREQ_001',
            'test_case_description': 'Verify House Hunt Request form is visible on home page',
            'test_steps': '1. Navigate to home page\n2. Locate House Hunt Request form',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'House Hunt Request form is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_HHREQ_001',
            'test_scenario_description': 'House Hunt Request Form',
            'test_case_id': 'TC_HHREQ_002',
            'test_case_description': 'Verify Submit Request button is present',
            'test_steps': '1. Navigate to home page\n2. Locate "Submit Request" button in the form',
            'preconditions': 'User is on the home page with request form',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Submit Request button is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_HHREQ_001',
            'test_scenario_description': 'House Hunt Request Form',
            'test_case_id': 'TC_HHREQ_003',
            'test_case_description': 'Verify Unit Type dropdown in form',
            'test_steps': '1. Navigate to home page\n2. Find Unit Type dropdown in the form\n3. Verify options are available',
            'preconditions': 'User is on the home page with request form',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Unit Type dropdown with options like "1 Bedroom" is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_HHREQ_001',
            'test_scenario_description': 'House Hunt Request Form',
            'test_case_id': 'TC_HHREQ_004',
            'test_case_description': 'Verify Budget dropdown in form',
            'test_steps': '1. Navigate to home page\n2. Find Budget dropdown in the form\n3. Verify options are available',
            'preconditions': 'User is on the home page with request form',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Budget dropdown with options like "Below 10k" is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_HHREQ_001',
            'test_scenario_description': 'House Hunt Request Form',
            'test_case_id': 'TC_HHREQ_005',
            'test_case_description': 'Verify location text field in form',
            'test_steps': '1. Navigate to home page\n2. Find location text field (placeholder: "e.g Kilimani, Westlands")',
            'preconditions': 'User is on the home page with request form',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Location text field is visible and editable',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_HHREQ_001',
            'test_scenario_description': 'House Hunt Request Form',
            'test_case_id': 'TC_HHREQ_006',
            'test_case_description': 'Verify form submission with valid data',
            'test_steps': '1. Navigate to home page\n2. Select Unit Type: "1 Bedroom"\n3. Select Budget: "Below 10k"\n4. Enter location: "nyeri"\n5. Click Submit Request',
            'preconditions': 'User is on the home page with request form',
            'test_data': 'Unit Type: 1 Bedroom, Budget: Below 10k, Location:yeri',
            'post_conditions': 'Form is submitted',
            'expected_result': 'Form submits successfully or shows confirmation',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 7. NEIGHBORHOOD SCORES & COMMUNITY TESTS
    # ==============================================================================
    
    neighborhood_test_cases = [
        {
            'test_scenario_id': 'TS_NB_001',
            'test_scenario_description': 'Neighborhood Scores & Ratings',
            'test_case_id': 'TC_NB_001',
            'test_case_description': 'Verify trending neighborhood scores section is displayed',
            'test_steps': '1. Navigate to home page\n2. Scroll to Trending Mtaa Scores section',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Trending neighborhood scores are displayed with ratings',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_NB_001',
            'test_scenario_description': 'Neighborhood Scores & Ratings',
            'test_case_id': 'TC_NB_002',
            'test_case_description': 'Verify neighborhood search functionality',
            'test_steps': '1. Navigate to home page\n2. Click on a neighborhood from trending scores\n3. Verify navigation to neighborhood details',
            'preconditions': 'User is on the home page with neighborhood scores',
            'test_data': 'None',
            'post_conditions': 'User is navigated to neighborhood details',
            'expected_result': 'Neighborhood page loads with scores and reviews',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_NB_002',
            'test_scenario_description': 'Community Watch Section',
            'test_case_id': 'TC_NB_003',
            'test_case_description': 'Verify community watch section is displayed',
            'test_steps': '1. Navigate to home page\n2. Scroll to "The Community Pulse" section',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Community watch section with live updates is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_NB_002',
            'test_scenario_description': 'Community Watch Section',
            'test_case_id': 'TC_NB_004',
            'test_case_description': 'Verify View All Alerts link in community section',
            'test_steps': '1. Navigate to home page\n2. Find "View All Alerts" link in community section\n3. Click on it',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User is navigated to community alerts page',
            'expected_result': 'Link navigates to /community page',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 8. TOP AGENTS & FEATURED PROPERTIES TESTS
    # ==============================================================================
    
    agents_test_cases = [
        {
            'test_scenario_id': 'TS_AGENT_001',
            'test_scenario_description': 'Top Agents Section',
            'test_case_id': 'TC_AGENT_001',
            'test_case_description': 'Verify Top Agents section is displayed on home page',
            'test_steps': '1. Navigate to home page\n2. Scroll to Top Agents section',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Top Agents section is visible with agent listings',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_AGENT_001',
            'test_scenario_description': 'Top Agents Section',
            'test_case_id': 'TC_AGENT_002',
            'test_case_description': 'Verify clicking on an agent navigates to profile',
            'test_steps': '1. Navigate to home page\n2. Click on a top agent card',
            'preconditions': 'User is on the home page with top agents',
            'test_data': 'None',
            'post_conditions': 'User is navigated to agent profile',
            'expected_result': 'Agent profile page is displayed',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 9. FAQ SECTION TESTS
    # ==============================================================================
    
    faq_test_cases = [
        {
            'test_scenario_id': 'TS_FAQ_001',
            'test_scenario_description': 'FAQ Section',
            'test_case_id': 'TC_FAQ_001',
            'test_case_description': 'Verify FAQ section is displayed on home page',
            'test_steps': '1. Navigate to home page\n2. Scroll to FAQ section\n3. Look for "Curious Minds Ask..." heading',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'FAQ section with questions and answers is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 10. STATS SECTION TESTS
    # ==============================================================================
    
    stats_test_cases = [
        {
            'test_scenario_id': 'TS_STATS_001',
            'test_scenario_description': 'Live Stats Section',
            'test_case_id': 'TC_STATS_001',
            'test_case_description': 'Verify animated stats section displays',
            'test_steps': '1. Navigate to home page\n2. Look for Live Stats section\n3. Check for "Real-time activity tracking" text',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Live Stats section with activity tracking is displayed',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_STATS_001',
            'test_scenario_description': 'Live Stats Section',
            'test_case_id': 'TC_STATS_002',
            'test_case_description': 'Verify stats show Total and Fulfilled numbers',
            'test_steps': '1. Navigate to home page\n2. Find Live Stats section\n3. Look for Total and Fulfilled values',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Stats show Total (e.g., 2495) and Fulfilled (e.g., 2445) counts',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 11. DECISION TOOLS TESTS
    # ==============================================================================
    
    decision_tools_test_cases = [
        {
            'test_scenario_id': 'TS_TOOL_001',
            'test_scenario_description': 'Neighbourhood Matchmaker AI',
            'test_case_id': 'TC_TOOL_001',
            'test_case_description': 'Verify Neighbourhood Matchmaker AI option is displayed',
            'test_steps': '1. Navigate to home page\n2. Scroll to "Make Smarter Decisions" section\n3. Look for Neighbourhood Matchmaker AI card',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'Neighbourhood Matchmaker AI option with "Launch Quiz" button is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_TOOL_001',
            'test_scenario_description': 'Neighbourhood Matchmaker AI',
            'test_case_id': 'TC_TOOL_002',
            'test_case_description': 'Verify clicking Launch Quiz navigates to quiz page',
            'test_steps': '1. Navigate to home page\n2. Click "Launch Quiz" button in Neighbourhood Matchmaker AI',
            'preconditions': 'User is on the home page with decision tools',
            'test_data': 'None',
            'post_conditions': 'User is navigated to find-my-neighbourhood page',
            'expected_result': 'Quiz page loads',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_TOOL_002',
            'test_scenario_description': 'Cost of Living Calculator',
            'test_case_id': 'TC_TOOL_003',
            'test_case_description': 'Verify Cost of Living Calculator option is displayed',
            'test_steps': '1. Navigate to home page\n2. Scroll to "Make Smarter Decisions" section\n3. Look for True Cost of Living card',
            'preconditions': 'User is on the home page',
            'test_data': 'None',
            'post_conditions': 'User remains on home page',
            'expected_result': 'True Cost of Living option with "Calculate Now" button is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_TOOL_002',
            'test_scenario_description': 'Cost of Living Calculator',
            'test_case_id': 'TC_TOOL_004',
            'test_case_description': 'Verify clicking Calculate Now navigates to calculator',
            'test_steps': '1. Navigate to home page\n2. Click "Calculate Now" button in True Cost of Living',
            'preconditions': 'User is on the home page with decision tools',
            'test_data': 'None',
            'post_conditions': 'User is navigated to cost-of-living calculator page',
            'expected_result': 'Calculator page loads',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # 12. NAVIGATION & HEADER/FOOTER TESTS
    # ==============================================================================
    
    navigation_test_cases = [
        {
            'test_scenario_id': 'TS_NAV_001',
            'test_scenario_description': 'Header Navigation',
            'test_case_id': 'TC_NAV_001',
            'test_case_description': 'Verify header is present on all pages',
            'test_steps': '1. Navigate to any page\n2. Check for header presence',
            'preconditions': 'User is on any page of the application',
            'test_data': 'None',
            'post_conditions': 'User remains on current page',
            'expected_result': 'Header with navigation is visible on all pages',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_NAV_001',
            'test_scenario_description': 'Header Navigation',
            'test_case_id': 'TC_NAV_002',
            'test_case_description': 'Verify footer is present on all pages',
            'test_steps': '1. Navigate to any page\n2. Scroll to footer',
            'preconditions': 'User is on any page of the application',
            'test_data': 'None',
            'post_conditions': 'User remains on current page',
            'expected_result': 'Footer is visible on all pages',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_NAV_001',
            'test_scenario_description': 'Header Navigation',
            'test_case_id': 'TC_NAV_003',
            'test_case_description': 'Verify WhatsApp contact button is visible',
            'test_steps': '1. Navigate to any page\n2. Look for WhatsApp button',
            'preconditions': 'User is on any page of the application',
            'test_data': 'None',
            'post_conditions': 'User remains on current page',
            'expected_result': 'WhatsApp contact button is visible',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
        {
            'test_scenario_id': 'TS_NAV_001',
            'test_scenario_description': 'Header Navigation',
            'test_case_id': 'TC_NAV_004',
            'test_case_description': 'Verify chat bubble is present',
            'test_steps': '1. Navigate to any page\n2. Look for chat bubble',
            'preconditions': 'User is on any page of the application',
            'test_data': 'None',
            'post_conditions': 'User remains on current page',
            'expected_result': 'Chat bubble is visible for support',
            'actual_result': '',
            'status': 'Not Executed',
            'executed_by': '',
            'executed_date': '',
            'comments': ''
        },
    ]
    
    # ==============================================================================
    # COMBINE ALL TEST CASES
    # ==============================================================================
    
    all_test_cases = (
        login_test_cases +
        registration_test_cases +
        home_page_test_cases +
        property_search_test_cases +
        property_listing_test_cases +
        house_hunt_request_test_cases +
        neighborhood_test_cases +
        agents_test_cases +
        faq_test_cases +
        stats_test_cases +
        decision_tools_test_cases +
        navigation_test_cases
    )
    
    # Add all test cases to the Excel file
    for test_case in all_test_cases:
        add_test_case(filename, test_case)
    
    print(f"\nAdded {len(all_test_cases)} test cases to the Excel file!")
    print(f"  - Login Tests: {len(login_test_cases)}")
    print(f"  - Registration Tests: {len(registration_test_cases)}")
    print(f"  - Home Page Tests: {len(home_page_test_cases)}")
    print(f"  - Property Search Tests: {len(property_search_test_cases)}")
    print(f"  - Property Listing Tests: {len(property_listing_test_cases)}")
    print(f"  - House Hunt Request Tests: {len(house_hunt_request_test_cases)}")
    print(f"  - Neighborhood Tests: {len(neighborhood_test_cases)}")
    print(f"  - Agents Tests: {len(agents_test_cases)}")
    print(f"  - FAQ Tests: {len(faq_test_cases)}")
    print(f"  - Stats Tests: {len(stats_test_cases)}")
    print(f"  - Decision Tools Tests: {len(decision_tools_test_cases)}")
    print(f"  - Navigation Tests: {len(navigation_test_cases)}")


if __name__ == "__main__":
    # Create the Excel file with headers
    filename = "Manual_Test_Documentation.xlsx"
    create_manual_test_excel(filename)
    
    # Add comprehensive test cases for House Hunt application
    create_house_hunt_test_cases(filename)
    
    print("\n" + "="*60)
    print("Manual Test Documentation Excel file created successfully!")
    print(f"File: {filename}")
    print("="*60)
